import ollama
import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from collections import defaultdict
from langchain_community.document_loaders import PyPDFLoader
from langchain_ollama import OllamaEmbeddings
from langchain_chroma import Chroma
from langchain.text_splitter import RecursiveCharacterTextSplitter

# ------------------ 教材來源 ------------------

PDF_LIST = [
    "Abraham Silberschatz Operating System Concepts.pdf",
    "Computer Organization and Design.pdf"
]

DB_DIR = "./db"
DB_INFO_FILE = os.path.join(DB_DIR, "dbinfo.json")

# ------------------ 檢查是否有新教材 ------------------

# 讀取之前已經匯入的教材檔名清單
def get_loaded_pdf_list():
    if os.path.exists(DB_INFO_FILE):
        with open(DB_INFO_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

# 儲存目前教材清單，以便下次比對是否有新增或異動
def save_loaded_pdf_list(pdf_list):
    os.makedirs(DB_DIR, exist_ok=True)
    with open(DB_INFO_FILE, "w", encoding="utf-8") as f:
        json.dump(pdf_list, f, ensure_ascii=False, indent=2)

# ------------------ 建立知識庫 ------------------

# 比對教材清單，確認是否有異動需要更新知識庫
need_refresh = sorted(PDF_LIST) != sorted(get_loaded_pdf_list())

if need_refresh:
    print("🧠 偵測到有新教材變動，開始重新建立向量資料庫...")
    documents = []

    for pdf_file in PDF_LIST:
        if os.path.exists(pdf_file):
            print(f"📘 載入教材：{pdf_file}")
            loader = PyPDFLoader(pdf_file)
            loaded_docs = loader.load()
            for doc in loaded_docs:
                doc.metadata["source_file"] = pdf_file  # 加入來源檔名以便查詢
            documents.extend(loaded_docs)
        else:
            print(f"⚠️ 找不到教材：{pdf_file}")

    # 分割段落內容供向量索引使用
    splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
    docs = splitter.split_documents(documents)

    # 建立向量資料庫
    embedding = OllamaEmbeddings(model="nomic-embed-text")
    vectordb = Chroma.from_documents(docs, embedding, persist_directory=DB_DIR)
    save_loaded_pdf_list(PDF_LIST)

    # 📚 產生章節摘要
    def generate_textbook_outline(docs):
        outlines = defaultdict(list)
        for doc in docs:
            source = doc.metadata.get("source_file", "未知教材")
            content = doc.page_content
            lines = content.splitlines()
            for line in lines:
                line = line.strip()
                if line.lower().startswith("chapter") or line[:3].isdigit():
                    if len(line) < 100:
                        outlines[source].append(line)
        return outlines

    outlines = generate_textbook_outline(documents)
    print("\n📚 教材章節目錄摘要：")
    for src, chapters in outlines.items():
        print(f"\n📘 {src}")
        for ch in chapters[:10]:
            print(f"  - {ch}")
else:
    print("✅ 教材未變動，直接載入資料庫")
    embedding = OllamaEmbeddings(model="nomic-embed-text")
    vectordb = Chroma(persist_directory=DB_DIR, embedding_function=embedding)

# 建立檢索器
retriever = vectordb.as_retriever()

# ------------------ 儲存對話紀錄 ------------------

# 儲存對話為 Excel，避免中文亂碼
def save_chat_xlsx(user_msg, bot_reply):
    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"chat_log_{today_str}.xlsx"
    time_str = datetime.now().strftime("%H:%M:%S")

    try:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["時間", "使用者", "AI 回覆"])

        ws.append([time_str, user_msg, bot_reply])
        wb.save(filename)
    except PermissionError:
        print("⚠️ 檔案可能已開啟中，請關閉 Excel 後重試。")

# ------------------ LLM 對話處理 ------------------

# 初始 prompt 訊息
messages = [{"role": "system", "content": "你是一位用繁體中文回答的家教老師，請用簡單方式講解問題。"}]
MODEL_NAME = 'llama3'
last_user = ""
last_reply = ""

# 問答主邏輯函式

def chat_with_ollama(prompt):
    global last_user, last_reply

    # 嘗試從 prompt 判斷是否有指定查詢某本教材
    selected_pdf = None
    for pdf_file in PDF_LIST:
        base = os.path.splitext(os.path.basename(pdf_file))[0].lower()
        if base in prompt.lower():
            selected_pdf = pdf_file
            break

    # 執行檢索，並根據指定教材過濾
    if selected_pdf:
        print(f"🎯 只搜尋教材：{selected_pdf}")
        all_docs = retriever.invoke(prompt)
        related_docs = [doc for doc in all_docs if doc.metadata.get("source_file") == selected_pdf][:3]
        if not related_docs:
            related_docs = all_docs[:3]  # 若指定結果太少則退回全部
    else:
        related_docs = retriever.invoke(prompt)[:3]

    # 建立上下文內容（含來源與頁數）
    context_chunks = []
    for doc in related_docs:
        source = doc.metadata.get("source_file", "未知教材")
        page = doc.metadata.get("page", "未知頁數")
        chunk = f"[來自：{source} 第 {page} 頁]\n{doc.page_content}"
        context_chunks.append(chunk)

    context_text = "\n---\n".join(context_chunks)

    # 包裝 prompt 給模型
    rag_prompt = f"""以下是教材內容摘要，請根據這些內容來回答問題：\n\n{context_text}\n\n使用者問題：{prompt}\n請用繁體中文詳細解釋，並舉例子說明。"""

    messages.append({"role": "user", "content": rag_prompt})

    try:
        response = ollama.chat(model=MODEL_NAME, messages=messages)
        reply = response['message']['content']
        messages.append({"role": "assistant", "content": reply})

        # 儲存對話紀錄
        save_chat_xlsx(prompt, reply)

        # 記住這次對話
        last_user = prompt
        last_reply = reply

        # 控制訊息堆積上限
        if len(messages) > 20:
            messages.pop(1)

        return reply
    except Exception as e:
        return f"❌ 發生錯誤：{e}"

# ------------------ 主程式 ------------------

if __name__ == "__main__":
    print("📘 家教系統已啟動，輸入 quit,exit,bye 離開")

    while True:
        user_input = input("你：").strip()

        if not user_input:
            continue

        if user_input.lower() in ["quit", "exit", "bye"]:
            print("👋 再見，祝學習愉快！")
            break

        response = chat_with_ollama(user_input)
        print("家教老師：", response)
