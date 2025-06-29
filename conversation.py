import ollama
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
from google import genai

def save_chat_xlsx(user_msg, bot_reply, folder_path="logs/chat"):
    # 確保資料夾存在，不存在就建立
    os.makedirs(folder_path, exist_ok=True)

    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"chat_log_{today_str}.xlsx"
    filepath = os.path.join(folder_path, filename)  # 完整路徑
    time_str = datetime.now().strftime("%H:%M:%S")

    try:
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["時間", "使用者", "AI 回覆"])

        ws.append([time_str, user_msg, bot_reply])
        wb.save(filepath)
    except PermissionError:
        print("⚠️ Excel 檔案正在開啟中，請關閉後再試")

def chat_with_ollama(prompt,messages,MODEL_NAME):
    try:
        response = ollama.chat(model=MODEL_NAME, messages=messages)
        reply = response['message']['content']
        messages.append({"role": "assistant", "content": reply})
        save_chat_xlsx(prompt, reply)
        if len(messages) > 20:
            messages.pop(1)
        return reply
    except Exception as e:
        return f"❌ 發生錯誤：{e}"

def chat_with_gemini(prompt,messages):
    try:
        load_dotenv()  # 自動讀取當前目錄下的 .env 檔
        api_key = os.getenv('GEMINI_API_KEY')
        client = genai.Client()

        response = client.models.generate_content(
            model="gemini-2.5-flash", contents = prompt
        )
        print(response.text)
        # response = ollama.chat(model=MODEL_NAME, messages=messages)
        # reply = response['message']['content']
        # messages.append({"role": "assistant", "content": reply})
        # save_chat_xlsx(prompt, reply)
        # if len(messages) > 20:
        #     messages.pop(1)
        # return reply
    except Exception as e:
        return f"❌ 發生錯誤：{e}"