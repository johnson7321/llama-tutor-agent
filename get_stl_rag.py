import ollama
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from langchain.docstore.document import Document
from langchain_ollama import OllamaEmbeddings
from langchain_chroma import Chroma
from langchain.text_splitter import RecursiveCharacterTextSplitter
from youtube_transcript_api import YouTubeTranscriptApi, NoTranscriptFound, TranscriptsDisabled
import re
import prompt
import shutil

# ------------------ 設定 ------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_DIR = os.path.join(SCRIPT_DIR, "yt_db")

# ------------------ 輸入影片連結 ------------------
def extract_youtube_id(url):
    pattern = r"(?:v=|\/)([0-9A-Za-z_-]{11})(?:\?|&|$)"
    match = re.search(pattern, url)
    return match.group(1) if match else None

url = input("請輸入 YouTube 網址：")
# url = "https://www.youtube.com/watch?v=TSfTqRCy4rc"
video_id = extract_youtube_id(url)

if not video_id:
    print("❌ 無法解析影片 ID")
    exit()

# ------------------ 擷取字幕並轉為 Document ------------------
try:
    transcript = YouTubeTranscriptApi.get_transcript(
        video_id, languages=['zh-TW', 'zh-Hant', 'zh-Hans', 'zh', 'en']
    )
    subtitles = [entry["text"] for entry in transcript]
    print(subtitles)
    full_text = "\n".join(subtitles)
    print("\n\nfull_text:"+full_text)
    yt_doc = Document(page_content=full_text, metadata={"source": f"YouTube_{video_id}"})
    documents = [yt_doc]
except NoTranscriptFound:
    print("⚠️ 無字幕")
    exit()
except TranscriptsDisabled:
    print("⚠️ 字幕功能被禁用")
    exit()
except Exception as e:
    print(f"❌ 發生錯誤：{e}")
    exit()

# ------------------ 建立向量資料庫 ------------------
# 顯示字幕預覽（前20行）
subtitles = [entry["text"] for entry in transcript]
#print("\n📝 字幕預覽（前20行）：")
# for i, line in enumerate(subtitles[:20]):
#     print(f"{i+1:02d}. {line}")

# 詢問是否繼續
choice = input("\n是否使用這些字幕建立知識庫？(y/n): ").strip().lower()
if choice != "y":
    print("❌ 已取消建立向量資料庫")
    exit()
    
# 建立向量資料庫前，清空舊資料庫
if os.path.exists(DB_DIR):
    shutil.rmtree(DB_DIR)

# ✅ 若確認，建立文件與向量資料庫
full_text = "\n".join(subtitles)
yt_doc = Document(page_content=full_text, metadata={"source": f"YouTube_{video_id}"})
documents = [yt_doc]

# 建立向量資料庫
os.makedirs(DB_DIR, exist_ok=True)
splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
docs = splitter.split_documents(documents)
embedding = OllamaEmbeddings(model="nomic-embed-text")
vectordb = Chroma.from_documents(docs, embedding, persist_directory=DB_DIR)
retriever = vectordb.as_retriever()
# ------------------ 對話邏輯 ------------------
messages = [{"role": "system", "content": prompt.tutor_guideline}]
MODEL_NAME = 'llama3.1:latest'

def save_chat_xlsx(user_msg, bot_reply, folder_path="logs/chat"):
    # 確保資料夾存在，不存在就建立
    os.makedirs(folder_path, exist_ok=True)

    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"chat_log_{today_str}.xlsx"
    filepath = os.path.join(folder_path, filename)  # 完整路徑
    time_str = datetime.now().strftime("%H:%M:%S")

    try:
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["時間", "使用者", "AI 回覆"])

        ws.append([time_str, user_msg, bot_reply])
        wb.save(filepath)
    except PermissionError:
        print("⚠️ Excel 檔案正在開啟中，請關閉後再試")

def chat_with_ollama(prompt):
    related_docs = retriever.invoke(prompt)[:3]
    context_chunks = [
        f"[來自：{doc.metadata.get('source', '未知')}]\n{doc.page_content}"
        for doc in related_docs
    ]
    context_text = "\n---\n".join(context_chunks)

    rag_prompt = f"""以下是影片字幕內容摘要，請根據這些內容來回答問題：\n\n{context_text}\n\n使用者問題：{prompt}\n請用繁體中文詳細解釋，並舉例子說明。"""

    messages.append({"role": "user", "content": rag_prompt})
    try:
        response = ollama.chat(model=MODEL_NAME, messages=messages)
        reply = response['message']['content']
        messages.append({"role": "assistant", "content": reply})
        save_chat_xlsx(prompt, reply)
        if len(messages) > 20:
            messages.pop(1)
        return reply
    except Exception as e:
        return f"❌ 發生錯誤：{e}"

# ------------------ 主程式 ------------------
print("🎥 字幕系統已啟動，輸入 quit,exit,bye 離開")

while True:
    user_input = input("你：").strip()
    if user_input.lower() in ["quit", "exit", "bye"]:
        print("👋 再見，祝學習愉快！")
        break
    if not user_input:
        continue
    response = chat_with_ollama(user_input)
    print("家教老師：", response)
